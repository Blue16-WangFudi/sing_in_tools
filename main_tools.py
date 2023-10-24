#encoding=utf-8
#author: blue16（Index:blue16.cn）
#date: 2023-10-24
#summary: 自动对比表单数据与人员名单，找出未签到和签退人员，并输出txt文本信息
#命名规则：
#10-24-A表示签到表格；10-24-B表示签退表格；namelist表示人员名单

import openpyxl

#默认不读取表头
def get_column(xlsx_path,id_column):
    #打开文件和对应工作簿
    coursebook=openpyxl.load_workbook(xlsx_path)
    coursebook_sheet=coursebook['Sheet1']
    #定义一个list存放结果
    ret=[]
    #获取最大行数
    max_row=coursebook_sheet.max_row
    for i in range(2,max_row+1):
        cell_id=id_column+str(i)
        if(coursebook_sheet[cell_id].value!=None):
            ret.append(coursebook_sheet[cell_id].value)
    return ret


#返回人员信息：[[name1,class1],[name2,class2]]
def get_memberinfo(xlsx_path,id_column_name,id_column_className,className):
    #打开文件和对应工作簿
    coursebook=openpyxl.load_workbook(xlsx_path)
    coursebook_sheet=coursebook['Sheet1']
    #定义一个list存放结果
    ret=[]
    #获取最大行数
    max_row=coursebook_sheet.max_row
    for i in range(2,max_row+1):
        temp=[]
        cell_id1=id_column_name+str(i)
        cell_id2=id_column_className+str(i)
        if(className!="*" and coursebook_sheet[cell_id2].value!=className):
            continue
        if(coursebook_sheet[cell_id1].value!=None):
            temp.append(coursebook_sheet[cell_id1].value)
        if(coursebook_sheet[cell_id2].value!=None):
            temp.append(coursebook_sheet[cell_id2].value)
        ret.append(temp)
    return ret
#删掉已经有了的人
def compare_delete(member_source,member_data):
    
    #temp1=一个名字
    for temp1 in member_data:
        count=0
        for temp2 in member_source:
             #找到了
            if(temp1==member_source[count][0]):
                member_source.pop(count)
            count=count+1

    return member_source

#按照班级分类：[[className,member1,member2,……],[className,member1,member2,……]，……]
def classify(list_name):
    list_className=[]
    #结构：[[className,member1,member2,……],[className,member1,member2,……]，……]
    list_classify=[]
    #temp2:一个[name,className]
    for temp2 in list_name:
        a=temp2[1] in list_className
        if(a == False):
            list_className.append(temp2[1])
            list_classify.append([temp2[1]])
        count=0
        #找到了，证明list_classify中有，那就直接在list_classify中找到对应项，添加进去
        #temp3=[className,member1,member2,……]
        for temp3 in list_classify:
            #该人员所在班级被找到
            if(temp3[0]==temp2[1]):
               temp3.append(temp2[0])
               #list_classify.pop(count)
               #list_classify.append(temp3)
            count=count+1
    return list_classify


def output(file_name,list_classify):
    temp=file_name.split(".")
    filename=temp[0]
    #三段了：月、日、A/B
    temp=temp[0].split("-")
    msg=""
    if(temp[2]=="A"):
        msg=str(temp[0])+"月"+str(temp[1])+"日未签到人员\n"
    if(temp[2]=="B"):
        msg=str(temp[0])+"月"+str(temp[1])+"日未签退人员\n"
    for temp in list_classify:
        first=True
        msg=msg+temp[0]+":"
        for temp2 in temp:
            if(first):
                #跳过第一个
                first=False
                continue
            msg=msg+temp2+"，"
        msg=msg+"\n"

    file=open(filename+".txt","w")
    file.write(msg)
    file.close()

#参数调整提示：这里设置名单信息
list_member=get_memberinfo("namelist.xlsx","A","C","*")
#参数调整提示：这里设置收集的数据信息
list_sign_in=get_column("10-24-A.xlsx","D")
list_result=compare_delete(list_member,list_sign_in)
list_classify=classify(list_result)
print(list_classify)
#参数调整提示：请注意输出文件名的写法
output("10-24-A.xlsx",list_classify)

